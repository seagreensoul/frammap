# -*- coding: utf-8 -*-

import openpyxl
import re
import geocoder
import folium

def popup_html(s1, t):
    html = '<style>b {font-size: 16pt;}table {font-size: 14pt;} </style>'
    html += '<h5><b>{}</b></h5>'.format(s1)
    for j in t:
        html += '<br><b>{}</b>'.format(ws[j][1].value)
        html += '<table border=1><tr><th>Г/М</th>'
        i=3
        html += '<th>Январь</th><th>Февраль</th><th>Март</th><th>Апрель</th><th>Май</th><th>Июнь</th><th>Июль</th><th>Август</th><th>Сентябрь</th><th>Октябрь</th><th>Ноябрь</th><th>Декабрь</th>'
        html += '</tr><tr><th>2016</th>'
        while i<26:
            try:
                html += '<td>{}</td>'.format(ws[j][i].value)
                i=i+2
            except IndexError:
                break
        i=4
        html += '</tr><tr><th>2017</th>'
        while i<28:
            try:
                html += '<td>{}</td>'.format(ws[j][i].value)
                i=i+2
            except IndexError:
                break
        html +='</table>'
    return html


wb = openpyxl.load_workbook(filename = 'C:/Users/admin/Downloads/Client Sales 2017_по территориям_Декабрь.xlsx',data_only = True)
ws = wb.get_sheet_by_name('ВАО')

map1 = folium.Map(location=[55.764414, 37.647859], zoom_start = 11)            
fg=folium.FeatureGroup(name="v")

i=3
while i<2050:
    j=i
    t=list()
    t.append(i)
    stri = ws[i][0].value
    while stri==ws[j+1][0].value:
        j= j+1
        t.append(j)

    s1 = re.sub(r'\([^\)]+\)', '', stri)
    s = re.findall(r"\((.*)\)", stri)
    if s[0].find('г.')==-1:
        s[0]='Москва '+ s[0]
    print (s[0])
    i=j+1
    try:
        g = geocoder.yandex(s[0])
        l=g.latlng
        print (l)
        a=float(l[0])
        b=float(l[1])
        fg.add_child(folium.Marker(location = [a,b], popup =  folium.Popup(popup_html(s1, t),max_width='auto') ))

        map1.add_child(fg)

    except TypeError:
        continue
    
map1.save(outfile="vao.html")